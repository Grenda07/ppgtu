# Run this app with `python app.py` and
# visit http://127.0.0.1:8050/ in your web browser.
import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
import pandas as pd
from datetime import datetime
import datetime
import base64
import locale
import graficos
import zipfile
import io
import matplotlib.pyplot as plt
from plotly.offline import plot
import numpy as np
import plotly.offline as offline
import plotly.graph_objs as go
import plotly.express as px
import tempfile
import os
import calendar
import graficos
import openpyxl
from openpyxl import workbook 
from openpyxl import load_workbook
from flask import Flask, send_file, make_response
import requests
import colorlover as cl

def mensal_bar(mes,tipo,ano,link,df): #ex: 1,Presencial
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if tipo   != 'todos':
		df     = df[df['tipo']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
		gray_palette = cl.scales['9']['seq']['Greys']
		fig = px.bar(df_sum, x='Hora', y='GRUPO', color='SUBCATEGORIA', orientation='h')
		fig.update_layout(title='Horas total trabalhadas por categoria e subcategoria')
		fig.update_layout( xaxis_title='Horas',yaxis_title='Categoria',legend_title='Subcategoria')
		if link == 1:
			fig=offline.plot(fig,output_type='div')
		return fig
	
def diario_bar (dia,tipo,df): #ex: 10/01/2022,Remoto
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	dia        = datetime.datetime.strptime(dia, '%d/%m/%Y')
	df         = df[df['DATA']== dia]
	if tipo   != 'todos':
		df     = df[df['tipo']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
		fig = px.bar(df_sum, x='Hora', y='GRUPO', color='SUBCATEGORIA', orientation='h')
		fig.update_layout( xaxis_title='Horas',yaxis_title='Categoria',legend_title='Subcategoria')
		fig.update_layout(title='Horas total trabalhadas no dia por categoria e subcategoria')
		return fig
	
	
def mensal_line(mes,tipo,ano,link,df): #ex: 1,Remoto
	year       = ano
	month      = mes
	start_date = pd.Timestamp(year, month, 1)
	end_date   = start_date + pd.offsets.MonthEnd(0)
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if tipo   != 'todos':
		df     = df[df['tipo']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','DATA']).agg({'Hora': 'sum'}).reset_index()
		color_map = {group: px.colors.qualitative.Plotly[i % len(px.colors.qualitative.Plotly)] for i, group in enumerate(df_sum['GRUPO'].unique())}
		colors = [color_map[group] for group in df_sum['GRUPO']]
		fig = px.bar(df_sum, x='DATA', y='Hora', color='GRUPO', orientation='v')
		#fig = go.Figure(data=[go.Scatter(x=df_sum[df_sum['GRUPO']==group]['DATA'], y=df_sum[df_sum['GRUPO']==group]['Hora'], mode='markers', marker=dict(color=color_map[group]), name=group) for group in df_sum['GRUPO'].unique()])
		fig.update_layout(xaxis_title='Data', yaxis_title='Horas', legend_title='Grupo')
		fig.update_layout(xaxis_range=[start_date,end_date])
		fig.update_layout(title='Horas trabalhadas por dia')
		fig.update_layout(xaxis_tickmode='linear')
		fig.update_layout(xaxis_tickangle=-90)
		if link ==1:
			fig=offline.plot(fig,output_type='div')
		return fig

def mensal_todos(mes,ano,link,df): #ex: 1
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['tipo']).agg({'Hora': 'sum'}).reset_index()
		fig        = px.bar(df_sum, x='tipo', y='Hora', color='tipo', orientation='v')
		fig.update_layout(yaxis_title='Horas',xaxis_title=' ',legend_title='Tipo')
		fig.update_layout(title='Horas total trabalhadas por tipo')
		if link ==1:
			fig=offline.plot(fig,output_type='div')
		return fig
        


def convert_to_time(decimal_num):
    hours = int(decimal_num)
    minutes = int(round((decimal_num - hours) * 60))
    return f"{hours:02d}:{minutes:02d}"



def preenche_modelo(mes,ano,nome,df): #ex: 1,Presencial
	url = 'https://github.com/Grenda07/ppgtu/blob/main/src/modelo.xlsx?raw=true'
	response = requests.get(url)
	content = response.content
	file = io.BytesIO(content)
	grupos=['Grupo de Pesquisa','Programa']
	tipo=['Presencial','Remoto']
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
	month_name = datetime.date(2000, mes, 1).strftime('%B')
	month_name = month_name.capitalize()
	meses = {"January": "Janeiro",
	         "February": "Fevereiro",
	         "March": "Mar??o",
	         "April": "Abril",
	         "May": "Maio",
	         "June": "Junho",
	         "July": "Julho",
	         "August": "Agosto",
	         "September": "Setembro",
	         "October": "Outubro",
	         "November": "Novembro",
	         "December": "Dezembro"}
	wb = load_workbook(file)
	sheets = wb.sheetnames
	Sheet1 = wb[sheets[0]] ##
	Sheet1.cell(row = 3, column = 1).value = nome
	Sheet1.cell(row = 3, column = 2).value = meses[month_name]
	Sheet1.cell(row = 3, column = 3).value = ano
	total= df['Hora'].sum() 
	Sheet1.cell(row = 10, column = 2).value = convert_to_time(total)
	Sheet1.cell(row = 7, column = 2).value = convert_to_time(total)
	df_sum     = df.groupby(['GRUPO']).agg({'Hora': 'sum'}).reset_index()
	df_sum2    = df.groupby(['tipo']).agg({'Hora': 'sum'}).reset_index()
	for index, (n, i) in enumerate(zip(grupos, tipo)):
		horas = df_sum.loc[df_sum['GRUPO'] == n, 'Hora'].iloc[0]
		Sheet1.cell(row = 8 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 8 + index, column = 3).value = horas*100/total
		horas = df_sum2.loc[df_sum2['tipo'] == i, 'Hora'].iloc[0]
		Sheet1.cell(row = 5 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 5 + index, column = 3).value = horas*100/total
	df_sum3    = df.groupby(['SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
	for index, row in df_sum3.iterrows():
		Sheet1.cell(row = 12 + index, column = 1).value = row['SUBCATEGORIA']
		Sheet1.cell(row = 12 + index, column = 2).value = convert_to_time(row['Hora'])
		df4    = df[df['SUBCATEGORIA']==row['SUBCATEGORIA']]
		Sheet1.cell(row = 12 + index, column = 3).value = df4['ATIVIDADE'].nunique()
	Sheet1.title = 'Relatorio'
	df = df.drop(['Hora'], axis=1)
	df['DATA'] = df['DATA'].dt.strftime('%d/%m/%Y') 
	df['HORAS'] = df['HORAS'].dt.strftime('%H:%M')
	
	df.to_excel('temp.xlsx', sheet_name='Dados detalhados',index=False)
	temp_wb = openpyxl.load_workbook('temp.xlsx')
	ws2 = wb.create_sheet("Dados detalhados")
	for row in temp_wb['Dados detalhados']:
		for cell in row:
			ws2[cell.coordinate].value = cell.value
	new_file_name = 'relatorio_'+meses[month_name]+'.xlsx'
	wb.save(new_file_name)
	return new_file_name
	
def retorna_df(contents, filename):
	contents=str(contents[0])
	content_type, content_string = contents.split(',')
	decoded = base64.b64decode(content_string)
	df = pd.read_excel(io.BytesIO(decoded))
	return df


# Initialize the app
app = dash.Dash(__name__)
server=app.server

# Define the layout
app.layout = html.Div([
    html.H1('Relat??rio Atividades PPGTU'),
    html.Div([dcc.Upload(id='upload-data',children=html.Div(['Arraste e solte ou ',html.A('selecione arquivos')]),
              style={'width': '50%','height': '60px','lineHeight': '60px','borderWidth': '1px',
                     'borderStyle': 'dashed','borderRadius': '5px','textAlign': 'center','margin': '10px'},
              multiple=True),
              html.Button('Submeter', id='transform-button'),
              html.Div(id='output-data-upload')]),
    html.Div(id='tipo-container', children=[
    dcc.RadioItems(
        id='freq-tipo',
        options=[
            {'label': 'Presencial', 'value': 'Presencial'},
            {'label': 'Remoto', 'value': 'Remoto'},
            {'label': 'Presencial e Remoto', 'value': 'todos'},
            {'label': 'Relat??rio' , 'value': 'relatorio'}
        ], value=None
    )]),
    html.Div(id='rela-container', children=[
    dcc.RadioItems(
        id='freq-radio',
        options=[
            {'label': 'Mensal', 'value': 'mensal'},
            {'label': 'Di??rio', 'value': 'diario'}
        ], value=None
    )], style={'display': 'none'}),  
    html.Div(id='mensal-container', children=[
        html.Label('Ano:'),
        html.Div([
        dcc.Input(id='year-input', type='number', placeholder='Ano')]),
        html.Div([
        html.Label('M??s:'),
        dcc.Dropdown(
            id='month-dropdown',
            options=[
                {'label': 'Janeiro'  , 'value': '01'},
                {'label': 'Fevereiro', 'value': '02'},
                {'label': 'Mar??o'    , 'value': '03'},
                {'label': 'Abril'    , 'value': '04'},
                {'label': 'Maio'     , 'value': '05'},
                {'label': 'Junho'    , 'value': '06'},
                {'label': 'Julho'    , 'value': '07'},
                {'label': 'Agosto'   , 'value': '08'},
                {'label': 'Setembro' , 'value': '09'},
                {'label': 'Outubro'  , 'value': '10'},
                {'label': 'Novembro' , 'value': '11'},
                {'label': 'Dezembro' , 'value': '12'}
            ],
            placeholder='M??s'
        )]),
        html.Button('Enter', id='submit-btn')
    ], style={'display': 'none'}),
    
    html.Div(id='diario-container', children=[
        dcc.Input(id='date-input', type='text', placeholder='DD/MM/YYYY'),
        html.Button('Enter', id='submit-btn-2')
    ], style={'display': 'none'}),
    
    html.Div(id="mensal-graphs1", children=[
        dcc.Graph(id="graph-1-mes"),
        dcc.Graph(id="graph-2")
    ], style={'display': 'none'}),
    
    html.Div(id="mensal-graphs2", children=[
        dcc.Graph(id="graph-1-1-mes"),
        dcc.Graph(id="graph-2-2"),
        dcc.Graph(id="graph-3")
    ], style={'display': 'none'}),
    
    html.Div(id="diario-graphs", children=[
        dcc.Graph(id="graph-1-dia")
    ], style={'display': 'none'}),
    
    html.Div(id="relatorio-container",children=[html.H2('Baixar'), 
        html.Label('Ano:'),
        html.Div([
        dcc.Input(id='year-input2', type='number', placeholder='Ano')]),
        html.Div([
        html.Label('M??s:'),
        dcc.Dropdown(
            id='month-dropdown2',
            options=[
                {'label': 'Janeiro'  , 'value': '01'},
                {'label': 'Fevereiro', 'value': '02'},
                {'label': 'Mar??o'    , 'value': '03'},
                {'label': 'Abril'    , 'value': '04'},
                {'label': 'Maio'     , 'value': '05'},
                {'label': 'Junho'    , 'value': '06'},
                {'label': 'Julho'    , 'value': '07'},
                {'label': 'Agosto'   , 'value': '08'},
                {'label': 'Setembro' , 'value': '09'},
                {'label': 'Outubro'  , 'value': '10'},
                {'label': 'Novembro' , 'value': '11'},
                {'label': 'Dezembro' , 'value': '12'}
            ],
            placeholder='M??s'
        )]),html.Label('Aluno(a):'),
                                  dcc.Input(id='name-input', type='text', placeholder='Digite seu nome'),
                                  html.Br(),
                                  html.Button('Download', id='download-link'),
                                  dcc.Download(id='download')],style={'display': 'none'})
])


# Define the callbacks
@app.callback(
    [Output('output-data-upload', 'children'),Output('mensal-container', 'style'), Output('diario-container', 'style'),Output("mensal-graphs1", "style"),Output("mensal-graphs2", "style"),
    Output("diario-graphs", "style"),Output("relatorio-container", "style"),Output("rela-container", "style")],
    [Input('freq-radio', 'value'),Input('freq-tipo', 'value'),Input('transform-button', 'n_clicks')]
)
def show_hide_divs(frequency,tipo,n_clicks):
	if not n_clicks:
		mensal_style         = {'display': 'none'}
		diario_style         = {'display': 'none'}
		mensal_graphs1_style = {'display': 'none'}
		mensal_graphs2_style = {'display': 'none'}
		diario_graphs_style  = {'display': 'none'}
		gerar_style          = {'display': 'none'}
		tipo_style           = {'display': 'none'}
		return {},mensal_style, diario_style, mensal_graphs1_style, mensal_graphs2_style, diario_graphs_style,gerar_style,tipo_style
	mensal_style         = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" else {'display': 'none'}
	diario_style         = {'display': 'block'} if tipo      != 'relatorio' and frequency == "diario" else {'display': 'none'}
	mensal_graphs1_style = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" and tipo != 'todos' else {'display': 'none'}
	mensal_graphs2_style = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" and tipo == 'todos' else {'display': 'none'}
	diario_graphs_style  = {'display': 'block'} if tipo      != 'relatorio' and frequency == "diario" else {'display': 'none'}
	gerar_style          = {'display': 'block'} if tipo      == 'relatorio' else {'display': 'none'}
	tipo_style           = {'display': 'block'} if tipo      == 'todos' or tipo == 'Remoto' or tipo == 'Presencial' else {'display': 'none'}
	return {},mensal_style, diario_style, mensal_graphs1_style, mensal_graphs2_style, diario_graphs_style,gerar_style,tipo_style

## PARA MENSAL (sem todos):
@app.callback(
    [Output('graph-1-mes', 'figure'), Output('graph-2', 'figure')],
    [Input('submit-btn', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('year-input', 'value'), State('month-dropdown', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_1(n_clicks,tipo, year, month,contents, filename):
	if not year or not month:
		return {}, {}
	df = retorna_df(contents, filename)
	fig1=mensal_bar(int(month),tipo,year,0,df)
	fig2=mensal_line(int(month),tipo,year,0,df)
	
	if fig1 == 'nan' or fig2 == 'nan':
		return {},{}
	else:
		return fig1, fig2

#PARA MENSAL (com todos):
@app.callback(
    [Output('graph-1-1-mes', 'figure'), Output('graph-2-2', 'figure'), Output('graph-3', 'figure')],
    [Input('submit-btn', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('year-input', 'value'), State('month-dropdown', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_2(n_clicks,tipo, year, month,contents, filename):
	if not year or not month:
		return {}, {},{}
	df = retorna_df(contents, filename)
	fig1=mensal_bar(int(month),tipo,year,0,df)
	fig2=mensal_line(int(month),tipo,year,0,df)
	fig3=mensal_todos(int(month),year,0,df)
	if fig1 == 'nan' or fig2 == 'nan' or fig3 == 'nan':
		return {},{},{}
	else:
		return fig1, fig2,fig3

#PARA RELAT??RIO:
@app.callback(
    Output('download', 'data'),
    [Input('download-link', 'n_clicks'),Input('name-input', 'value')],
    [State('year-input2', 'value'), State('month-dropdown2', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_2(n_clicks,nome, year, month,contents, filename):
	if not year or not month:
		return None
	df = retorna_df(contents, filename)
	if n_clicks is not None:
		return dcc.send_file(preenche_modelo(int(month),year,nome,df))
	else:
		return None

#PARA DI??RIO:
@app.callback(
    Output('graph-1-dia', 'figure'),
    [Input('submit-btn-2', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('date-input', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_3(n_clicks,tipo, date,contents, filename):
	if not date:
		return {}
	df = retorna_df(contents, filename)
	fig = diario_bar (date,tipo,df)
	if fig == 'nan':
		return {}
	else:
		return fig

if __name__ == '__main__':
    app.run_server(debug=True)
